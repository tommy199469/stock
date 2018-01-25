import pandas as pd
import requests , datetime , os , xlsxwriter
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# get data from html
def get_data():
    page = requests.get('http://www.etnet.com.hk/www/tc/stocks/industry_adu.php')
    soup = BeautifulSoup(page.text, 'lxml')
    rows = soup.find_all( "tr" , {"class": ["oddRow", "evenRow"]})
    plate = []

    for i in rows:
        plate.append(
            {
                "name" : i.find('td').getText() ,
                "平均升/跌幅" : i.findAll('td')[1].getText()
            }
        )
    return plate


def analysis_data(data , excel_name):
    sheet_name_yestaday = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime("%B %d, %Y")
    sheet_name_5days    = (datetime.datetime.now() - datetime.timedelta(days=5)).strftime("%B %d, %Y")
    sheet_name_7days    = (datetime.datetime.now() - datetime.timedelta(days=7)).strftime("%B %d, %Y")
    writer = pd.ExcelWriter(excel_name, engine='openpyxl')
    book = None
    if os.path.isfile(excel_name):
        book = load_workbook(excel_name)
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    else:
        return False

    today_data = []
    yestaday_data = []
    five_data = []
    seven_data = []

    for item in data:
        today_data.append(item['平均升/跌幅'].replace("%", ""))

    if sheet_name_yestaday in writer.sheets:
        sheet = book[sheet_name_yestaday]
        cells = sheet['B2': 'C34']
        for c1, c2 in cells:
            yestaday_data.append(c2.value.replace("%", ""))

    if sheet_name_5days in writer.sheets:
        worksheet = writer.sheets[sheet_name_5days]

    if sheet_name_7days in writer.sheets:
        worksheet = writer.sheets[sheet_name_7days]


    for index ,i in enumerate(today_data):
        if yestaday_data:
            data[index]['比較一日前'] = str(round((float(i) + (float(yestaday_data[index])) ) , 2))+"%"

        if five_data:
            data[index]['比較五日前'] = str(round((float(i) + (float(five_data[index])) ) , 2))+"%"

        if seven_data:
            data[index]['比較七日前'] = str(round((float(i) + (float(seven_data[index])) ) , 2))+"%"

    return data


# hanlde on excel
def handle_excel(data , excel_name):
    sheet_name = datetime.date.today().strftime("%B %d, %Y")
    try:
        df = pd.DataFrame(data)
        writer = pd.ExcelWriter(excel_name, engine='openpyxl')
        if os.path.isfile(excel_name):
            book = load_workbook(excel_name)
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        df.to_excel(writer, sheet_name=sheet_name)
        column_list = pd.Series(range(len(df.columns))).tolist()
        worksheet = writer.sheets[sheet_name]
        for column_cells  in column_list:
            i = get_column_letter(column_cells+2)
            worksheet.column_dimensions[i].width = 20

        writer.save()
        writer.close()
    except Exception as e:
        print(e)
        return False

    return True
